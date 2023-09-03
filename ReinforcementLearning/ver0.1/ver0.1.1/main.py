import numpy as np
from CrewPairingEnv import CrewPairingEnv
import pandas as pd
from cal import reshape_list, calculateScore
from getData import getInputData
from PolicyGradientAgent import PolicyGradientAgent

initial_pairing_set = pd.read_excel(
    '/home/public/yunairline/ASCP/ReinforcementLearning/dataset/output.xlsx', header=1)
del initial_pairing_set['pairing index']

n_pairings = len(initial_pairing_set)
max_flight = len(initial_pairing_set.columns)
column = list(range(1, max_flight+1))
column_list = [str(item) for item in column]
initial_pairing_set.columns = column_list
initial_pairing_set = initial_pairing_set.fillna(-1)
pairing_matrix = initial_pairing_set.values  # <class 'numpy.ndarray'> (215, 4)

env = CrewPairingEnv(pairing_matrix)
agent = PolicyGradientAgent(env)

input_flight, input_deadhead, input_salary = getInputData()

"""
하나의 episode에서 하는 일
1. 랜덤하게 flihgt를 선택해서 state를 받음
2. step 진행하여 갱신이 있을 때까지 반복
3. 받은 reward를 저장
4. step이 끝나면 저장한 reward를 통해 policy를 업데이트
"""

for i_episode in range(10):
    print('episode : ', i_episode+1)

    # state reset하기
    state_history = []
    # 리셋하는 부분이 꼭 필요함.!!! -> 여기서 변경된 페어링 샛을 기준으로 새로운 state를 생성할 수 있도록 하고 episode가 끝나면 다시 리셋
    state = env.reset()
    state_history.append(state)  # 한번 선택한 비행은 다시 선택하지 않도록 state_history에 저장
    # print('pairing_set : ', pairing_set, 'state : ', state,state_history)
    done = False

    # 하나의 step에서 하는 일
    # 1. state를 받아서 action을 선택
    # 2. 선택한 action을 통해 다음 state와 reward를 받음

    for t in range(150):
        print('step : ', t+1)
        state, action = agent.select_action(state)
        # available action이 없는 경우 action가 None이 됨 -> 이 경우에는 step을 종료
        if action == None:
            print('No available actions.')
            break

        # idx를 통해서 flight 정보 출력하기
        state = int(state)
        action = int(action)
        print('before_flight : ',
              env.pairing_set[state], 'after_flight : ', env.pairing_set[action])

        # step 진행 -> 진행하면 env.pairing_set이 갱신됨 -> 안되게 하려면 env.pairing_set을 복사해서 사용해야 함
        input_pairing_set = env.pairing_set.copy()
        step_pairing_set, reward, done, _ = env.step(
            state, action, input_pairing_set)
        print('done : ', done, 'reward : ', reward,
              'is step_pairing_set None? : ', step_pairing_set is None, end='\n\n')
        agent.rewards.append(reward)
        # pairingset은 done이 True인 경우 갱신된 pairingset으로 바뀜 False인 경우에는 그냥 처음의 pairingset을 유지
        if step_pairing_set is not None:
            env.pairing_set = step_pairing_set
            # state_history.append(action)
        else:
            pass
        # print('pairing_set : ', env.pairing_set)

    agent.update()

    reshaped_list = reshape_list(
        env.pairing_set, num_rows=215, num_columns=4)

    hard_score = 0
    soft_score = 0
    for pair in reshaped_list:
        pair_hard_score, pair_soft_score = calculateScore(pair)
        hard_score += pair_hard_score
        soft_score += pair_soft_score

    # print('hard score : ', hard_score, 'soft score : ', soft_score)

    # 'pairing index'를 포함한 데이터프레임 생성
    df = pd.DataFrame(reshaped_list)

    # 'pairing index' 열 추가
    df.insert(0, 'pairing index', range(len(df)))

    # -1 값을 빈 값으로 변경
    df = df.replace(-1.0, '')

    col_num = len(df.columns)
    score_list = ['soft', soft_score, 'hard', hard_score]
    while len(score_list) < col_num:
        score_list.extend([None])

    # 맨 위 행에 score값 추가
    df.loc[-1] = score_list
    df.index = df.index + 1
    df = df.sort_index()

    # 칼럼명과 첫번째 행 바꾸기
    tobe_col = df.iloc[0]
    tobe_first = df.columns.tolist()
    tobe_first = [np.nan if isinstance(
        x, (int, float)) else x for x in tobe_first]

    df.columns = tobe_col
    df.loc[0] = tobe_first

    import openpyxl
    # 엑셀 파일로 저장
    file_path = "/home/public/yunairline/ASCP/ReinforcementLearning/output/output.xlsx"
    # 파일의 이름을 확인하고 없으면 새로운 파일 생성
    try:
        workbook = openpyxl.load_workbook(file_path)
    except:
        workbook = openpyxl.Workbook()
        workbook.save(file_path)
    workbook = openpyxl.load_workbook(file_path)

    # 새로운 시트 생성
    new_sheet_name = 'episode'+str(i_episode+1)
    new_sheet = workbook.create_sheet(title=new_sheet_name)

    df.to_excel('/home/public/yunairline/ASCP/ReinforcementLearning/output/output' +
                str(new_sheet_name)+'.xlsx', sheet_name='episode'+str(i_episode+1), index=False)
