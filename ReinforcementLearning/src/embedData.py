"""
데이터를 읽어오고, vector shape를 구성하기 위한 함수

input: xlsx file
output: vector shape -> [T_origin, T_dest, T_dur, A_origin, A_dest]

example: [int, int, int, [0,1,...,0], [1,0,...,0]]
"""

import sys
import pandas as pd
import openpyxl
from IDProvider import IDProvider
from Flight import Flight
from Components import Aircraft, Airport, Hotel
import torch
import os 
import random
from sklearn.cluster import KMeans
def readXlsx(path, inputFileName):

    print(path)

    xls = pd.ExcelFile(path + inputFileName)  # 엑셀 파일의 모든 시트 읽기
    sheet_names = xls.sheet_names  # 엑셀 파일 안의 각 시트 이름을 리스트형태로 가져옴

    for sheet_name in sheet_names:  # 각 시트를 순회하며 csv파일로 저장
        df = pd.read_excel(path + inputFileName, sheet_name=sheet_name)
        csv_filename = f"{path}/{sheet_name}.csv"
        df.to_csv(csv_filename, index=False)

# 공항 총 리스트
def airportList(A_origin, A_dest):
    # 공항의 총 개수 파악 -> 두 열의 합집합
    A_total = list(set(A_origin) | set(A_dest))
    # 알파벳 순서로 정렬
    A_total.sort()
    # 순서대로 onehot으로 mapping
    return A_total

def aircraftList(Aircraft):
    # 항공기의 총 개수 파악 -> 두 열의 합집합
    Aircraft_total = list(set(Aircraft))
    # 알파벳 순서로 정렬
    Aircraft_total.sort()
    # 순서대로 onehot으로 mapping
    return Aircraft_total

def embedFlightData(path): # flight 객체 생성 및 vector로 변환, flight_list, V_f_list 반환
    idprovider = IDProvider() 
    fdf = pd.read_csv(path+'/User_Flight.csv')
    # 200개 행 읽어오기
    #fdf = fdf.head(200)
    fdf = fdf.drop(fdf.index[0])
    fdf.columns = fdf.iloc[0]
    fdf = fdf.drop(fdf.index[0])
    flight_list=[] # flight 객체를 저장할 리스트
    V_f_list = [] # flight 객체를 vector로 변환하여 저장할 리스트

    for idx, row in fdf.iterrows(): # flight 객체 생성
        flight = Flight(
            idx=idprovider.get_flight_id(),  # 싱글톤인 idprovider 호출하여 flight에 id 부여
            TailNumber=row['T/N'],
            originAirport=row['ORIGIN'],
            originTime=row['ORIGIN_DATE'],
            destAirport=row['DEST'],
            destTime=row['DEST_DATE'],
            aircraft=row['AIRCRAFT_TYPE']
        )
        flight_list.append(flight)
    
    flight_list=sorted(flight_list) # originTime 기준으로 정렬(originTime이 같다면 destTime 기준으로 정렬)
    
    airport_total = airportList(fdf['ORIGIN'], fdf['DEST'])
    aircraft_total = aircraftList(fdf['AIRCRAFT_TYPE'])
    
    ddf=pd.read_csv(path+'/User_Deadhead.csv')
    ddf = ddf.drop(ddf.index[0])
    ddf.columns = ddf.iloc[0]
    ddf = ddf.drop(ddf.index[0])
    

    for idx, row in ddf.iterrows():
        airport_origin_onehot = [0 for _ in range(len(airport_total))]
        airport_dest_onehot = [0 for _ in range(len(airport_total))]

        for i, airport in enumerate(airport_total):
            if airport == row['출발 공항']:
                airport_origin_onehot[i] = 1
            if airport == row['도착 공항']:
                airport_dest_onehot[i] = 1

        ddf.at[idx, '출발 공항'] = airport_origin_onehot
        ddf.at[idx, '도착 공항'] = airport_dest_onehot
        Airport.add_edge(row['출발 공항'], row['도착 공항'], row['Deadhead(원)'])

    
    cdf=pd.read_csv(path+'/Program_Cost.csv')
    cdf.columns = cdf.iloc[0]
    cdf = cdf.drop(cdf.index[0])
    for idx,row in cdf.iterrows():
        aircraft_onehot = [0 for _ in range(len(aircraft_total))]
        
        for i,aircraft in enumerate(aircraft_total):
            if aircraft == row['AIRCRAFT']:
                aircraft_onehot[i] = 1
        cdf.at[idx, 'AIRCRAFT'] = aircraft_onehot
        Aircraft.add_type(row['AIRCRAFT'], row['CREW_NUM(명)'], int(float(row['Layover Cost(원/분)'])), int(float(row['Quick Turn Cost(원/회)'])))
    temp=tuple([0 for _ in range(len(aircraft_total))])
    del Aircraft.dic[temp]
    
    
    hdf=pd.read_csv(path+'/User_Hotel.csv')
    hdf = hdf.drop(hdf.index[0])
    hdf.columns = hdf.iloc[0]
    hdf=hdf.drop(hdf.index[0])
    hdf = hdf.dropna(subset=[hdf.columns[0]])
    for idx,row in hdf.iterrows():
        airport_onehot = [0 for i in range(len(airport_total))]
        
        for i,airport in enumerate(airport_total):
            if airport == row['공항 Code']:
                airport_onehot[i] = 1
        hdf.at[idx, '공항 Code'] = airport_onehot
        Hotel.add_hotel(row['공항 Code'], row['비용(원)'])
    
    for i in range(len(flight_list)):
        V_f_list.append(flight_list[i].toVector(airport_total, aircraft_total))


    # Neural net size : (시간 2진수 배열 -> 16bit)*2 + (공항 Onehot 배열 size)*2
    NN_size = (2 + 2*len(airport_total)) * 2

    return flight_list, V_f_list, NN_size
    
def embedFlightData_Random(path, sample_size=None): 
    idprovider = IDProvider() 
    fdf = pd.read_csv(path+'/User_Flight.csv')
    fdf = fdf.drop(fdf.index[0])
    fdf.columns = fdf.iloc[0]
    fdf = fdf.drop(fdf.index[0])
    
    # 랜덤 샘플링 적용
    if sample_size:
        fdf = fdf.sample(n=sample_size, random_state=42)

    flight_list=[] 
    V_f_list = [] 
    
    for idx, row in fdf.iterrows(): # flight 객체 생성
        flight = Flight(
            idx=idprovider.get_flight_id(),  # 싱글톤인 idprovider 호출하여 flight에 id 부여
            TailNumber=row['T/N'],
            originAirport=row['ORIGIN'],
            originTime=row['ORIGIN_DATE'],
            destAirport=row['DEST'],
            destTime=row['DEST_DATE'],
            aircraft=row['AIRCRAFT_TYPE']
        )
        flight_list.append(flight)
    
    flight_list=sorted(flight_list) # originTime 기준으로 정렬(originTime이 같다면 destTime 기준으로 정렬)
    
    airport_total = airportList(fdf['ORIGIN'], fdf['DEST'])
    aircraft_total = aircraftList(fdf['AIRCRAFT_TYPE'])
    
    ddf=pd.read_csv(path+'/User_Deadhead.csv')
    ddf = ddf.drop(ddf.index[0])
    ddf.columns = ddf.iloc[0]
    ddf = ddf.drop(ddf.index[0])
    

    for idx, row in ddf.iterrows():
        airport_origin_onehot = [0 for _ in range(len(airport_total))]
        airport_dest_onehot = [0 for _ in range(len(airport_total))]

        for i, airport in enumerate(airport_total):
            if airport == row['출발 공항']:
                airport_origin_onehot[i] = 1
            if airport == row['도착 공항']:
                airport_dest_onehot[i] = 1

        ddf.at[idx, '출발 공항'] = airport_origin_onehot
        ddf.at[idx, '도착 공항'] = airport_dest_onehot
        Airport.add_edge(row['출발 공항'], row['도착 공항'], row['Deadhead(원)'])

    
    cdf=pd.read_csv(path+'/Program_Cost.csv')
    cdf.columns = cdf.iloc[0]
    cdf = cdf.drop(cdf.index[0])
    for idx,row in cdf.iterrows():
        aircraft_onehot = [0 for _ in range(len(aircraft_total))]
        
        for i,aircraft in enumerate(aircraft_total):
            if aircraft == row['AIRCRAFT']:
                aircraft_onehot[i] = 1
        cdf.at[idx, 'AIRCRAFT'] = aircraft_onehot
        Aircraft.add_type(row['AIRCRAFT'], row['CREW_NUM(명)'], int(float(row['Layover Cost(원/분)'])), int(float(row['Quick Turn Cost(원/회)'])))
    temp=tuple([0 for _ in range(len(aircraft_total))])
    del Aircraft.dic[temp]
    
    
    hdf=pd.read_csv(path+'/User_Hotel.csv')
    hdf = hdf.drop(hdf.index[0])
    hdf.columns = hdf.iloc[0]
    hdf=hdf.drop(hdf.index[0])
    hdf = hdf.dropna(subset=[hdf.columns[0]])
    for idx,row in hdf.iterrows():
        airport_onehot = [0 for i in range(len(airport_total))]
        
        for i,airport in enumerate(airport_total):
            if airport == row['공항 Code']:
                airport_onehot[i] = 1
        hdf.at[idx, '공항 Code'] = airport_onehot
        Hotel.add_hotel(row['공항 Code'], row['비용(원)'])
    
    for i in range(len(flight_list)):
        V_f_list.append(flight_list[i].toVector(airport_total, aircraft_total))
    # Neural net size : (시간 2진수 배열 -> 16bit)*2 + (공항 Onehot 배열 size)*2
    NN_size = (2 + 2*len(airport_total)) * 2

    return flight_list, V_f_list, NN_size

def embedFlightData_Interval(path, interval=2): 
    idprovider = IDProvider() 
    fdf = pd.read_csv(path+'/User_Flight.csv')
    fdf = fdf.drop(fdf.index[0])
    fdf.columns = fdf.iloc[0]
    fdf = fdf.drop(fdf.index[0])
    # 시간순서대로 정렬
    fdf['ORIGIN_DATE'] = pd.to_datetime(fdf['ORIGIN_DATE'])
    fdf = fdf.sort_values(by='ORIGIN_DATE')

    # 시스템적 샘플링 적용
    fdf = systematicSampling(fdf, interval)

    
    flight_list=[] 
    V_f_list = [] 

    for idx, row in fdf.iterrows(): 
        flight = Flight(
            idx=idprovider.get_flight_id(),
            TailNumber=row['T/N'],
            originAirport=row['ORIGIN'],
            originTime=row['ORIGIN_DATE'],
            destAirport=row['DEST'],
            destTime=row['DEST_DATE'],
            aircraft=row['AIRCRAFT_TYPE']
        )
        flight_list.append(flight)
    
    flight_list = sorted(flight_list)
    
    airport_total = airportList(fdf['ORIGIN'], fdf['DEST'])
    aircraft_total = aircraftList(fdf['AIRCRAFT_TYPE'])
    
    ddf=pd.read_csv(path+'/User_Deadhead.csv')
    ddf = ddf.drop(ddf.index[0])
    ddf.columns = ddf.iloc[0]
    ddf = ddf.drop(ddf.index[0])
    

    for idx, row in ddf.iterrows():
        airport_origin_onehot = [0 for _ in range(len(airport_total))]
        airport_dest_onehot = [0 for _ in range(len(airport_total))]

        for i, airport in enumerate(airport_total):
            if airport == row['출발 공항']:
                airport_origin_onehot[i] = 1
            if airport == row['도착 공항']:
                airport_dest_onehot[i] = 1

        ddf.at[idx, '출발 공항'] = airport_origin_onehot
        ddf.at[idx, '도착 공항'] = airport_dest_onehot
        Airport.add_edge(row['출발 공항'], row['도착 공항'], row['Deadhead(원)'])

    
    cdf=pd.read_csv(path+'/Program_Cost.csv')
    cdf.columns = cdf.iloc[0]
    cdf = cdf.drop(cdf.index[0])
    for idx,row in cdf.iterrows():
        aircraft_onehot = [0 for _ in range(len(aircraft_total))]
        
        for i,aircraft in enumerate(aircraft_total):
            if aircraft == row['AIRCRAFT']:
                aircraft_onehot[i] = 1
        cdf.at[idx, 'AIRCRAFT'] = aircraft_onehot
        Aircraft.add_type(row['AIRCRAFT'], row['CREW_NUM(명)'], int(float(row['Layover Cost(원/분)'])), int(float(row['Quick Turn Cost(원/회)'])))
    temp=tuple([0 for _ in range(len(aircraft_total))])
    del Aircraft.dic[temp]
    
    
    hdf=pd.read_csv(path+'/User_Hotel.csv')
    hdf = hdf.drop(hdf.index[0])
    hdf.columns = hdf.iloc[0]
    hdf=hdf.drop(hdf.index[0])
    hdf = hdf.dropna(subset=[hdf.columns[0]])
    for idx,row in hdf.iterrows():
        airport_onehot = [0 for i in range(len(airport_total))]
        
        for i,airport in enumerate(airport_total):
            if airport == row['공항 Code']:
                airport_onehot[i] = 1
        hdf.at[idx, '공항 Code'] = airport_onehot
        Hotel.add_hotel(row['공항 Code'], row['비용(원)'])
    
    for i in range(len(flight_list)):
        V_f_list.append(flight_list[i].toVector(airport_total, aircraft_total))
        
    # Neural net size : (시간 2진수 배열 -> 16bit)*2 + (공항 Onehot 배열 size)*2
    NN_size = (2 + 2*len(airport_total)) * 2

    return flight_list, V_f_list, NN_size

def embedFlightData_Stratified(path, interval=2):
    idprovider = IDProvider()
    flight_list = []  # flight 객체를 저장할 리스트
    V_f_list = []  # flight 객체를 vector로 변환하여 저장할 리스트
    
    # Read User_Flight.csv
    fdf = pd.read_csv(path + '/User_Flight.csv')
    fdf = fdf.drop(fdf.index[0])
    fdf.columns = fdf.iloc[0]
    fdf = fdf.drop(fdf.index[0])
    
    # Drop unnecessary rows and columns
    fdf = fdf.iloc[1:, 1:]
    
    flight_list = [
        Flight(
            idx=idprovider.get_flight_id(),
            TailNumber=row['T/N'],
            originAirport=row['ORIGIN'],
            originTime=row['ORIGIN_DATE'],
            destAirport=row['DEST'],
            destTime=row['DEST_DATE'],
            aircraft=row['AIRCRAFT_TYPE']
        ) for idx, row in fdf.iterrows()
    ]

    # Sort flight objects by originTime and destTime
    flight_list.sort()

    # Filter and stratify the flight data
    sampled_data = stratifiedSampling(fdf, interval)

    airport_total = airportList(sampled_data['ORIGIN'], sampled_data['DEST'])
    aircraft_total = aircraftList(sampled_data['AIRCRAFT_TYPE'])

    # Read User_Deadhead.csv and create Airport edges
    ddf=pd.read_csv(path+'/User_Deadhead.csv')
    ddf = ddf.drop(ddf.index[0])
    ddf.columns = ddf.iloc[0]
    ddf = ddf.drop(ddf.index[0])
    for idx, row in ddf.iterrows():
        airport_origin_onehot = [1 if airport == row['출발 공항'] else 0 for airport in airport_total]
        airport_dest_onehot = [1 if airport == row['도착 공항'] else 0 for airport in airport_total]
        ddf.at[idx, '출발 공항'] = airport_origin_onehot
        ddf.at[idx, '도착 공항'] = airport_dest_onehot
        Airport.add_edge(row['출발 공항'], row['도착 공항'], row['Deadhead(원)'])

    # Read Program_Cost.csv and create Aircraft types
    cdf=pd.read_csv(path+'/Program_Cost.csv')
    cdf.columns = cdf.iloc[0]
    cdf = cdf.drop(cdf.index[0])
    for idx, row in cdf.iterrows():
        aircraft_onehot = [1 if aircraft == row['AIRCRAFT'] else 0 for aircraft in aircraft_total]
        cdf.at[idx, 'AIRCRAFT'] = aircraft_onehot
        Aircraft.add_type(row['AIRCRAFT'], row['CREW_NUM(명)'], int(float(row['Layover Cost(원/분)'])),
                           int(float(row['Quick Turn Cost(원/회)'])))

    temp = tuple([0 for _ in range(len(aircraft_total))])
    del Aircraft.dic[temp]

    # Read User_Hotel.csv and create Hotel data
    hdf=pd.read_csv(path+'/User_Hotel.csv')
    hdf = hdf.drop(hdf.index[0])
    hdf.columns = hdf.iloc[0]
    hdf=hdf.drop(hdf.index[0])
    hdf = hdf.dropna(subset=[hdf.columns[0]])
    for idx, row in hdf.iterrows():
        airport_onehot = [1 if airport == row['공항 Code'] else 0 for airport in airport_total]
        hdf.at[idx, '공항 Code'] = airport_onehot
        Hotel.add_hotel(row['공항 Code'], row['비용(원)'])

    flight_list = [
        Flight(
            idx=idprovider.get_flight_id(),
            TailNumber=row['T/N'],
            originAirport=row['ORIGIN'],
            originTime=row['ORIGIN_DATE'],
            destAirport=row['DEST'],
            destTime=row['DEST_DATE'],
            aircraft=row['AIRCRAFT_TYPE']
        ) for idx, row in sampled_data.iterrows()
    ]

    # Vectorize flight data
    V_f_list = [flight.toVector(airport_total, aircraft_total) for flight in flight_list]

    # Neural net size: (시간 2진수 배열 -> 16bit)*2 + (공항 Onehot 배열 size)*2
    NN_size = (2 + 2 * len(airport_total)) * 2

    return flight_list, V_f_list, NN_size

def stratifiedSampling(data, interval):
    # Group data by originAirport and destAirport
    grouped_data = data.groupby(['ORIGIN', 'DEST'])

    # Stratified sampling for each group
    sampled_data = pd.concat([systematicSampling(group_data, interval) for _, group_data in grouped_data])
    print(sampled_data)
    return sampled_data

def print_xlsx(output):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Pairing data 제목 추가
    sheet.cell(row=1, column=1, value="Pairing data")

    # 데이터를 엑셀에 쓰기
    for row_index, row_data in enumerate(output, start=2):  # 첫 번째 행은 이미 Pairing data로 사용되었으므로 2부터 시작
        # 각 행의 첫 열에는 1부터 시작하는 인덱스 추가
        sheet.cell(row=row_index, column=1, value=row_index - 1)

        # 나머지 데이터 추가
        for col_index, value in enumerate(row_data, start=2):  # 각 행의 두 번째 열부터 시작
            sheet.cell(row=row_index, column=col_index, value=value)

    workbook.save("output.xlsx")
    
def print_xlsx_tmp(n_epi,number,output_tmp,folder_path):
    if n_epi%number==0:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Pairing data 제목 추가
        sheet.cell(row=1, column=1, value="Pairing data")
        sheet.cell(row=1, column=2, value=f"episode_{n_epi}")

        # 데이터를 엑셀에 쓰기
        for row_index, row_data in enumerate(output_tmp, start=2):  # 첫 번째 행은 이미 Pairing data로 사용되었으므로 2부터 시작
            # 각 행의 첫 열에는 1부터 시작하는 인덱스 추가
            sheet.cell(row=row_index, column=1, value=row_index - 1)

            # 나머지 데이터 추가
            for col_index, value in enumerate(row_data, start=2):  # 각 행의 두 번째 열부터 시작
                sheet.cell(row=row_index, column=col_index, value=value)

        # 저장
        file_name = f"output_{n_epi}.xlsx"
        file_path = os.path.join(folder_path, file_name)

        workbook.save(file_path)
    

def flatten(V_p, V_f):
    V_p3 = V_p[3]
    V_p4 = V_p[4]
    
    if V_p[3] == [0] :
        V_p3 = [0] * len(V_f[3])
        V_p4 = [0] * len(V_f[4])

    nn_input = [V_p[0]//100000] + [V_p[1]//100000] + V_p3 + V_p4 + [V_f[0]//100000] + [V_f[1]//100000] + V_f[3] + V_f[4]

    return nn_input

def systematicSampling(data, interval):
    # 시작 인덱스 랜덤 선택
    start_index = random.randint(0, interval - 1)
    
    # 데이터프레임에서 샘플 추출
    sampled_data = data.iloc[start_index::interval]
    
    return sampled_data

